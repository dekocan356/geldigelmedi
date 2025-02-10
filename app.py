#!/usr/bin/env python3
import os
import time
import logging
import configparser
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
import openpyxl
from openpyxl import Workbook
from rapidfuzz import process, fuzz

# Flask uygulama ayarları
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
app.config['RESULT_FOLDER'] = os.path.join(os.getcwd(), 'results')
app.secret_key = 'gizli_anahtar'  # Güvenlik anahtarı

# Eğer klasörler yoksa oluşturun
for folder in [app.config['UPLOAD_FOLDER'], app.config['RESULT_FOLDER']]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Log ayarları
logging.basicConfig(
    filename="webgeldigelmedi.log",
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logging.info("Web uygulaması başlatıldı.")

@app.route('/', methods=['GET', 'POST'])
def upload_files():
    if request.method == 'POST':
        # Dosyaları ve form verilerini alıyoruz.
        adsoyad_file = request.files.get('adsoyad')
        kontrol_files = request.files.getlist('kontrol')
        sheet_name = request.form.get('sheet_name', 'Kontrol').strip()

        if not adsoyad_file or adsoyad_file.filename == "":
            flash("Lütfen Ad Soyad Listesi dosyasını yükleyin.")
            return redirect(request.url)
        if not kontrol_files or kontrol_files[0].filename == "":
            flash("Lütfen en az bir Kontrol Listesi dosyası yükleyin.")
            return redirect(request.url)

        # Dosyaları uploads klasörüne kaydediyoruz.
        adsoyad_path = os.path.join(app.config['UPLOAD_FOLDER'], adsoyad_file.filename)
        adsoyad_file.save(adsoyad_path)
        saved_kontrol_files = []
        for file in kontrol_files:
            if file.filename != "":
                path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                file.save(path)
                saved_kontrol_files.append(path)

        # Dosyaları işleyip Bulunamayanlar dosyasını oluşturuyoruz.
        result_file = process_files(adsoyad_path, saved_kontrol_files, sheet_name)
        flash("İşlem tamamlandı. 'Bulunamayanlar.xlsx' dosyasını indirin.")
        # Sonuç dosyasını results klasörüne kaydediyoruz.
        return redirect(url_for('download_file', filename=os.path.basename(result_file)))
    return render_template('upload.html')


def process_files(adsoyad_path, kontrol_paths, sheet_name):
    # İlk dosya: A sütunundaki isimleri oku.
    try:
        wb_adsoyad = openpyxl.load_workbook(adsoyad_path, data_only=True)
        ws_adsoyad = wb_adsoyad.active
    except Exception as e:
        logging.exception("Ad Soyad Listesi dosyası okunurken hata oluştu:\n%s", e)
        return None

    # İsimleri küçük harf haliyle bir sözlüğe kaydediyoruz.
    unmatched_dict = {}
    for row in ws_adsoyad.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            name = str(row[0]).strip()
            unmatched_dict[name.lower()] = row

    # Kontrol dosyalarında eşleşmeleri kontrol ediyoruz.
    for file in kontrol_paths:
        try:
            wb_control = openpyxl.load_workbook(file)
            if sheet_name and sheet_name in wb_control.sheetnames:
                ws_control = wb_control[sheet_name]
            else:
                ws_control = wb_control.active
        except Exception as e:
            logging.exception("Kontrol dosyası okunurken hata: %s", file)
            continue

        # B sütununda veri bulunan son satırı buluyoruz.
        max_row = 0
        for cell in ws_control["B"]:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
        if max_row == 0:
            continue

        # B sütunundaki her hücre için; eşleşme kontrolü yapıyoruz.
        for row in ws_control.iter_rows(min_row=1, max_row=max_row):
            if len(row) < 2:
                continue
            cell_b = row[1]
            if cell_b.value:
                cell_val = str(cell_b.value).strip().lower()
                if unmatched_dict:
                    # RapidFuzz eşik değeri: %80 (gerekirse burada değeri değiştirebilirsiniz, örn: 87)
                    result = process.extractOne(cell_val, list(unmatched_dict.keys()), scorer=fuzz.ratio)
                    if result and result[1] >= 80:
                        # Eşleşme sağlanırsa, G sütununa "geldi" yazıyoruz.
                        ws_control.cell(row=cell_b.row, column=7, value="geldi")
                        matched_name = result[0]
                        if matched_name in unmatched_dict:
                            del unmatched_dict[matched_name]
        # Kontrol dosyasını güncellenmiş haliyle kaydedelim.
        base, ext = os.path.splitext(file)
        new_file = base + "_guncellendi" + ext
        try:
            wb_control.save(new_file)
            logging.info("Güncellenmiş kontrol dosyası kaydedildi: %s", new_file)
        except Exception as e:
            logging.exception("Kontrol dosyası kaydedilirken hata: %s", e)

    logging.info("İşlem tamamlandı. Eşleşmeyen isimler: %s", list(unmatched_dict.keys()))

    # Bulunamayanlar dosyasını results klasörüne yazalım.
    output_path = os.path.join(app.config['RESULT_FOLDER'], "Bulunamayanlar.xlsx")
    wb_unmatched = Workbook(write_only=True)
    ws_unmatched = wb_unmatched.create_sheet("Bulunamayanlar")
    ws_unmatched.append(["İsim", "Orijinal Satır Verisi"])
    if unmatched_dict:
        for name, row_data in unmatched_dict.items():
            ws_unmatched.append([name] + list(row_data))
    else:
        ws_unmatched.append(["Tüm satırlar eşleşti."])
    try:
        wb_unmatched.save(output_path)
        logging.info("Bulunamayanlar dosyası oluşturuldu: %s", output_path)
    except Exception as e:
        logging.exception("Bulunamayanlar dosyası kaydedilirken hata oluştu:\n%s", e)
    return output_path


@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['RESULT_FOLDER'], filename, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
